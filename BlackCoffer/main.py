import openpyxl
import nltk
nltk.download('punkt')
nltk.download('vader_lexicon')
import requests
from bs4 import BeautifulSoup
from nltk.tokenize import sent_tokenize, word_tokenize
import re
import csv
from nltk.sentiment import SentimentIntensityAnalyzer
import time
# Load the Excel file
# wb = openpyxl.load_workbook('Output Data Structure.xlsx')
#
# # Select a specific worksheet
# sheet = wb['Sheet1'] # Replace 'Sheet1' with the actual sheet name
# col_list=[]

#
# # Loop will print all columns name
# for i in range(1, max_col + 1):
#     cell_obj = sheet.cell(row=1, column=i)
#     col_list.append(cell_obj.value)
# print(col_list)
def load_stop_words(stop_words_file):
    with open(stop_words_file, 'r') as file:
        stop_words = file.read().splitlines()
    return stop_words

def clean_text(text, stop_words):
    tokens = word_tokenize(text)
    cleaned_tokens = [token.lower() for token in tokens if token.lower() not in stop_words and token.isalpha()]
    cleaned_text = ' '.join(cleaned_tokens)
    return cleaned_text

# Creating a dictionary of Positive and Negative words
positive_words = set()
negative_words = set()
stop_words=set(load_stop_words('stop_words_file.txt'))
print(stop_words)
# Adding positive and negative words to the respective sets
with open('positive-words.txt', 'r') as file:
    for line in file:
        word = line.strip()
        if word not in stop_words:
            positive_words.add(word)

with open('negative-words.txt', 'r') as file:
    for line in file:
        word = line.strip()
        if word not in stop_words:
            negative_words.add(word)
# Extracting Derived variables
def calculate_sentiment(text):
    tokens = word_tokenize(text)
    positive_score = sum(1 for token in tokens if token in positive_words)
    negative_score = sum(1 for token in tokens if token in negative_words)
    polarity_score = (positive_score - negative_score) / ((positive_score + negative_score) + 0.000001)
    subjectivity_score = (positive_score + negative_score) / (len(tokens) + 0.000001)
    return [positive_score, negative_score, polarity_score, subjectivity_score]

# Analysis of Readability
def calculate_readability(text):
    sentences = sent_tokenize(text)
    num_sentences = len(sentences)
    num_words = 0
    num_complex_words = 0

    for sentence in sentences:
        words = word_tokenize(sentence)
        num_words += len(words)
        for word in words:
            
            if syllable_count(word) > 2:
                num_complex_words += 1

    avg_sentence_length = num_words / num_sentences
    percentage_complex_words = num_complex_words / num_words
    fog_index = 0.4 * (avg_sentence_length + percentage_complex_words)

    return [avg_sentence_length, percentage_complex_words, fog_index]

#Average Number of Words Per Sentence
def calculate_avg_words_per_sentence(text):
    sentences = sent_tokenize(text)
    num_sentences = len(sentences)
    num_words = 0

    for sentence in sentences:
        words = word_tokenize(sentence)
        num_words += len(words)

    avg_words_per_sentence = num_words / num_sentences
    return avg_words_per_sentence

#Complex Word Count
def calculate_complex_word_count(text):
    words = word_tokenize(text)
    complex_word_count = sum(1 for word in words if syllable_count(word) > 2)
    return complex_word_count

#Word Count
def calculate_word_count(text):
    words = word_tokenize(text)
    word_count = len(words)
    return word_count

#Syllable Count Per Word
def syllable_count(word):
    vowels = 'aeiou'
    syllables = 0
    prev_char_vowel = False

    for char in word:
        if char.lower() in vowels:
            if not prev_char_vowel:
                syllables += 1
            prev_char_vowel = True
        else:
            prev_char_vowel = False

    if word.endswith(('es', 'ed')):
        syllables -= 1

    return syllables

#Personal Pronouns
def count_personal_pronouns(text):
    personal_pronouns = ['I', 'we', 'my', 'ours', 'us']
    # Exclude the country name 'US'
    regex = r'\b(?![Uu][Ss]\b)\w+\b'
    matches = re.findall(regex, text)
    personal_pronoun_count = sum(1 for match in matches if match.lower() in personal_pronouns)
    return personal_pronoun_count

#Average Word Length
def calculate_avg_word_length(text):
    words = word_tokenize(text)
    total_char_count = sum(len(word) for word in words)
    avg_word_length = total_char_count / len(words)
    return avg_word_length

def main():
    wb = openpyxl.load_workbook('input.xlsx')
    sheet = wb['Sheet1']
    max_row = sheet.max_row
    # Access the URLs in a specific column
    url_column = 'URL'  # Replace 'URL' with the actual column name in your Excel file

    # Get the column index of the URL column

    url_column_index = None
    for cell in sheet[1]:
        if cell.value == url_column:
            url_column_index = cell.column_letter
            break
    result_list = [
        ['URL_ID', 'URL', 'POSITIVE SCORE', 'NEGATIVE SCORE', 'POLARITY SCORE', 'SUBJECTIVITY SCORE',
         'AVG SENTENCE LENGTH', 'PERCENTAGE OF COMPLEX WORDS', 'FOG INDEX', 'AVG NUMBER OF WORDS PER SENTENCE',
         'COMPLEX WORD COUNT', 'WORD COUNT', 'SYLLABLE PER WORD', 'PERSONAL PRONOUNS', 'AVG WORD LENGTH'],
    ]
    url_data = [cell.value for cell in sheet[url_column_index][1:]]

    # Print the URLs
    for i, url in enumerate(url_data):
        response = requests.get(url)

        # Parse the HTML content using Beautiful Soup
        soup = BeautifulSoup(response.text, 'html.parser')

        # Find all the paragraphs on the page
        paragraphs = soup.find_all('p')

        # Iterate over paragraphs and convert them into a list of sentences
        res_sen = ''

        for paragraph in paragraphs:
            paragraph_text = paragraph.get_text()
            res_sen += paragraph_text + " "
        cleaned_text = clean_text(res_sen, stop_words)
        sent=calculate_sentiment(cleaned_text)
        # print(f"sent list is: {sent}")
        readability_list=calculate_readability(cleaned_text)
        # print(f"read list is: {readability_list}")
        res=[sheet.cell(i+2,1).value,url]
        res.extend(sent)
        res.extend(readability_list)
        res.append(calculate_avg_words_per_sentence(res_sen))
        res.append(calculate_complex_word_count(res_sen))
        res.append(calculate_word_count(cleaned_text))
        res.append(syllable_count(res_sen))
        res.append(count_personal_pronouns(res_sen))
        res.append(calculate_avg_word_length(cleaned_text))
        print(res)
        result_list.append(res)
    print(result_list)
    csv_file = 'output.csv'

    # Write the data to the CSV file
    with open(csv_file, 'w', newline='') as file:
        writer = csv.writer(file)
        writer.writerows(result_list)

if __name__ == '__main__':
    main()